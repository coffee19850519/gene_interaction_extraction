import xlrd as xlrd
import openpyxl
import os
from label_file import LabelFile
from OCR import display
import cfg




def read_excel_data_by_name(file_name):
    table = None
    error_msg = None
    try:
        data = xlrd.open_workbook(file_name)
        table = data.sheet_by_index(0)
        # data.sheet_by_name(sheet_name)
    except Exception as msg:
        error_msg = msg
    return table, error_msg


def get_column_index(table, column_name):
    column_index = None

    # print table
    for i in range(table.ncols):
        # print columnName
        # print table.cell_value(0, i)
        if table.cell_value(0, i) == column_name:
            column_index = i
            break
    return column_index


def load_dictionary_from_excl(col_num):
    table, err = read_excel_data_by_name(cfg.dictionary_path)
    row_num = table.nrows
    words = []
    for row_idx in range(1, row_num):
        entity = table.cell_value(row_idx, col_num)
        entity = str(entity).upper()
        words.append(entity)
    return words


def load_ground_truth_into_excl(col_num):
    context_list = []
    dict_list = []

    for image_file in os.listdir(cfg.ground_truth_folder):
        image_name, image_ext = os.path.splitext(image_file)
        try:
            label_gt = LabelFile(os.path.join(cfg.ground_truth_folder, image_name + '.json'))
            for gt_idx in range(len(label_gt.shapes)):
                gt_category, gt_context = label_gt.shapes[gt_idx]['label'].split(':', 1)
                if gt_category == "gene" or gt_category == 'compound' or gt_category == 'location' \
                        or gt_category == 'ref_function' or gt_category == 'Title' or gt_category == 'other':
                    context_list.append(gt_context.strip())
        except Exception as err:
            print('Exception: ', image_name, err)
            continue

    if len(context_list) > 0:
        wb = openpyxl.Workbook()

        if cfg.previous_dictionary_path:  # previous dictionary, if needed or no ground truth
            wb = openpyxl.load_workbook(cfg.previous_dictionary_path)  # not tested

        ws = wb.active
        # ws = wb.get_sheet_by_name(sheet_name)

        ws.cell(column=col_num, row=1).value = "Name:"

        for row_num in range(2, ws.max_row + 1):
            dict_list.append(ws.cell(column=col_num, row=row_num).value.strip())  # get previous dictionary entries

        display("Loading ground truth into dictionary" + "\n", file=cfg.log_file)

        for context in context_list:

            context = context.upper().replace('\n', '')
            context = context.strip()

            display(str(context), file=cfg.word_file, to_print=False)

            max_row = ws.max_row
            if context and context not in dict_list:
                display(str(context), file=cfg.log_file)
                ws.cell(column=col_num, row=ws.max_row + 1).value = context
                dict_list.append(context)

        wb.save(cfg.dictionary_path)

        display("\n" + "Finished loading ground truth into dictionary" + "\n", file=cfg.log_file)



if __name__ == '__main__':
    filename = r'C:\Users\hefe\PycharmProjects\gene_interaction_extraction\gene_dictionary.xlsx'
    L = load_dictionary_from_excl(filename, 'Sheet1', 0)
    # table, err = read_excel_data_by_name(filename, 'Sheet1')
    # rowNum=table.nrows
    # l = []
    # for i in range(1, rowNum):
    #     gene = table.cell_value(i, 0)
    #     l.append(gene)
    print(L)

# end of file
