import pandas as pd
import codecs
import openpyxl
from openpyxl.utils import get_column_letter

def excel_one_line_to_list():
    df1 = pd.read_excel("dictionary.xlsx", usecols=[0], sheet_name='Sheet')  # 读取项目名称列,不要列名
    df_li1 = df1.values.tolist()
    result1 = []
    for s_li1 in df_li1:
        result1.append(s_li1[0])
    #result1=list(set(result1))
    #print(result1)

    df = pd.read_excel("text_result.xlsx", usecols=[0], sheet_name='Sheet1')  # 读取项目名称列,不要列名
    df_li = df.values.tolist()
    result2 = []
    for s_li in df_li:
        result2.append(s_li[0])
    result2 = list(set(result2))
    #print(result2)

    result = result1 + result2
    result = list(set(result))

    deal(result)


def deal(result):
    # 列表
    name_list = result

    # list转dataframe
    df = pd.DataFrame(name_list, columns=['Name:'])

    # 保存到本地excel
    df.to_excel("/Users/orion/Desktop/use_case1/predict/gene_dictionary.xlsx", index=False)

def txt_to_xlsx(filename,outfile):

    fr = codecs.open(filename,'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws = wb.create_sheet()
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row +=1
        line = line.strip()
        line = line.split('\t')
        col = 0
        for j in range(len(line)):
            col +=1
            #print (line[j])
            ws.cell(column = col,row = row,value = line[j].format(get_column_letter(col)))
    wb.save(outfile)

#if __name__ == '__main__':
#    excel_one_line_to_list()
